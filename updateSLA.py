import traceback
from datetime import datetime
import shutil
import os
import pandas as pd
from collections import Counter
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table
import win32com.client

# Instructions based on https://byu.instructure.com/courses/1026/pages/sla-update-tutorial

# Key variables
templatePath = r"C:\Users\wywyguy\Box\Accessibility Shared Folder\Team Member Folders\Wyatt\Programs\SLA Automation\SLA Report Template.xlsx"
mainReportPath = r"C:\Users\wywyguy\Box\Accessibility Shared Folder\Team Member Folders\Wyatt\Programs\SLA Automation\SLA Report Overview.xlsx"
logPath = r"C:\Users\wywyguy\Box\Accessibility Shared Folder\Team Member Folders\Wyatt\Programs\SLA Automation\SLA Update Program Log.txt"

# Function to log errors and warnings
def log(message):
    with open(logPath, "a", encoding="utf-8") as file:
        timestamp = datetime.now().strftime("%m/%d/%Y %H:%M:%S")
        if message == "BEGIN EXECUTION":
            file.write("\n\n\n")
        file.write(f"{timestamp}: {message}\n")

# Determine which data is which and rename files accordingly
def determineData(paths, now):
    keywords = {
        "prototype": "Prototypes",
        "50%": "50s",
        "psia": "PSIAs",
        "peer": "Peer Verifications"
    }

    fileLabels = {}
    labelCounts = Counter()
    undeterminedFiles = []

    for path in paths:
        if not os.path.exists(path):
            fileLabels[path] = None
            undeterminedFiles.append(path)
            continue

        try:
            df = pd.read_excel(path, engine='openpyxl')
            if 'I' not in df.columns and 8 not in df.columns:
                col_i = df.iloc[:, 8]
            else:
                col_i = df['I'] if 'I' in df.columns else df.iloc[:, 8]
            
            termCounter = Counter()
            for val in col_i.dropna().astype(str):
                for term in keywords:
                    if term.lower() in val.lower():
                        termCounter[term] += 1
            
            if termCounter:
                mostCommonTerm = termCounter.most_common(1)[0][0]
                label = keywords[mostCommonTerm]
                fileLabels[path] = label
                labelCounts[label] += 1
            else:
                fileLabels[path] = None
                undeterminedFiles.append(path)
            
        except Exception as e:
            print(f"Error reading {path}: {e}")
            fileLabels = None
            undeterminedFiles.append(path)

    knownLabels = [label for label in fileLabels.values() if label]
    if len(set(knownLabels)) != len(knownLabels):
        raise ValueError("Duplicate file types detected among existing files.")
    if len(fileLabels) != 4:
        raise ValueError("Total number of files (known + undetermined) must be 4.")

    yearMonth = now.strftime("%Y-%m")
    renamedFiles = {}

    for path, label in fileLabels.items():
        if label:
            newName = f"{yearMonth} - SLA {label}.xlsx"
            renamedFiles[path] = newName
            os.rename(path, os.path.join(os.path.dirname(path), newName))

    return renamedFiles, undeterminedFiles


# Core functionality
log("BEGIN EXECUTION")
try:
    # Step 1: Copy the SLA report template
    downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
    os.chdir(downloads_path)
    now = datetime.now()
    reportPath=f"{now:%Y-%m} - SLA Report.xlsx"
    shutil.copy2(templatePath, reportPath)

    # Step 2: Rename downloaded files according to contents of column I
    defaultPaths=[
        f"All Tasks Report - {now.strftime('%d %b %Y')}.xlsx",
        f"All Tasks Report - {now.strftime('%d %b %Y')} (1).xlsx",
        f"All Tasks Report - {now.strftime('%d %b %Y')} (2).xlsx",
        f"All Tasks Report - {now.strftime('%d %b %Y')} (3).xlsx"
    ]
    determineData(defaultPaths, now)

    # Step 3: Remove irregular tasks
    paths = {
        f"{now.strftime('%Y-%m')} - SLA Prototypes.xlsx": "Prototype Review - Accessibility",
        f"{now.strftime('%Y-%m')} - SLA 50s.xlsx": "50% Review - Accessibility",
        f"{now.strftime('%Y-%m')} - SLA PSIAs.xlsx": "Complete PSIA (Post-Supplier Inspection—Accessibility)",
        f"{now.strftime('%Y-%m')} - SLA Peer Verifications.xlsx": "Complete a Peer Verification"
    }
    for path, expected in paths.items():
        if os.path.exists(path):
            df = pd.read_excel(path, engine='openpyxl')
            col_i = df['I'] if 'I' in df.columns else df.iloc[:, 8]
            mask = col_i.dropna().astype(str).str.lower().str.contains(expected.lower(), regex=False, na=False)
            cleaned_df = df[mask]
            cleaned_df.to_excel(path, index=False, engine='openpyxl')
            if len(df)-len(cleaned_df) != 0:
                log(f"Removed {len(df)-len(cleaned_df)} entries from {path}.")
        else:
            log(f"Warning: File at '{path}' not found when trying to clean data. This indicates that this file was empty when downloaded or is missing.")

    # Step 4: Copy data to the monthly report
    copyMapping = {
        1: 'A',  # B -> A
        8: 'B',  # I -> B
        11: 'C', # L -> C
        14: 'D', # O -> D
        19: 'E', # T -> E
        23: 'F', # X -> F
        24: 'G', # Y -> G
        25: 'H'  # Z -> H
    }
    sheetMapping = {
        f"{now.strftime('%Y-%m')} - SLA Prototypes.xlsx": "Prototypes",
        f"{now.strftime('%Y-%m')} - SLA 50s.xlsx": "50% Reviews",
        f"{now.strftime('%Y-%m')} - SLA PSIAs.xlsx": "PSIAs",
        f"{now.strftime('%Y-%m')} - SLA Peer Verifications.xlsx": "Peer Verifications"
    }
    for path in paths:
        if os.path.exists(path):
            df = pd.read_excel(path, header=None, engine='openpyxl')
            df = df.iloc[:, list(copyMapping.keys())]
            df = df.dropna(how='all')
            df.columns = [copyMapping[i] for i in copyMapping]
            with pd.ExcelWriter(reportPath, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                df.to_excel(writer, sheet_name=sheetMapping[path], startrow=1, index=False, header=False)
        else:
            log(f"Warning: File at '{path}' not found when trying to copy data. This indicates that this file was empty when downloaded or is missing.")

    # Step 5: Clean the monthly report
    wb = load_workbook(reportPath)
    for sheet in sheetMapping.values():
        if sheet in wb.sheetnames:
            ws = wb[sheet]
            toDelete = []
            for row in range(2, ws.max_row + 1):
                value = ws[f"A{row}"].value
                if value is None or str(value).strip() == "":
                    toDelete.append(row)
            for row in reversed(toDelete):
                ws.delete_rows(row)
            for tbl in ws.tables.values():
                last_row = ws.max_row
                tbl.ref=f"A2:I{last_row}"
        else:
            log(f"Warning: Sheet {sheet} not found in SLA Report.")
    wb["Overview"]["B4"] = f"{now.strftime('%B %Y')} SLA Report Overview"
    wb.save(reportPath)

    # Step 6: Move the montly report to its designated location
    newDestination = f"N:\IS\Quality Assurance\ACCESSIBILITY\SLA Monthly Reports\{now.strftime('%Y')} SLA\{reportPath}"
    shutil.move(reportPath, newDestination)
    reportPath = newDestination
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    openReport = excel.Workbooks.Open(reportPath, UpdateLinks=1)
    openReport.Save()
    openReport.Close()
    excel.Quit()

    # Step 7: Add a new row in the main document for the month and link to the monthly report
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    mainReport = excel.Workbooks.Open(mainReportPath, UpdateLinks=1)
    mainSheet = mainReport.Sheets("SLA and Time Data")
    lastDataRow = mainSheet.Range("A2").End(-4121).Row
    mainSheet.Rows(lastDataRow).Insert()
    mainSheet.Cells(lastDataRow, 1).Value = now.strftime('%b-%y')
    sourceCells = [ # Cells from monthly report Overview page
        "C13", "C14", "C10",
        "F13", "F14", "F10",
        "C24", "C25", "C21",
        "F24", "F25", "F21"
    ]
    for i, cellRef in enumerate(sourceCells):
        targetCol = i + 2
        formula = f"='N:\IS\Quality Assurance\ACCESSIBILITY\SLA Monthly Reports\{now.strftime('%Y')} SLA\[{now.strftime('%Y-%m')} - SLA Report.xlsx]Overview'!${sourceCells[i][0]}${sourceCells[i][1:]}"
        mainSheet.Cells(lastDataRow, targetCol).Formula = formula
    mainReport.Save()
    mainReport.Close()
    excel.Quit()

    # Step 8: Remove the 4 default reports
    for path in paths:
        if os.path.exists(path):
            os.remove(path)

except Exception as e:
    log(f"ERROR: {e}")
    log(traceback.format_exc())